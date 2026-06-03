import re
import pdfplumber
import pandas as pd
from openpyxl import Workbook, load_workbook


class ConverterService:
    """Converts a PDF file to an Excel workbook.

    Strategy:
      - For each page, try to extract tables via pdfplumber.
      - If a page has no tables, fall back to raw text lines.
      - Each page becomes its own sheet in the workbook.
    """

    def analyze(self, pdf_path: str) -> dict:
        with pdfplumber.open(pdf_path) as pdf:
            page_count = len(pdf.pages)

        return {
            'page_count': page_count,
        }

    def convert(self, pdf_path: str, excel_path: str, progress_cb=None, single_sheet: bool = False) -> None:
        self.convert_range(pdf_path, excel_path, 1, None, progress_cb=progress_cb, single_sheet=single_sheet)

    def convert_range(self, pdf_path: str, excel_path: str, start_page: int, end_page: int | None, progress_cb=None, single_sheet: bool = False) -> None:
        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        single_ws = None
        invoice_header_written = False
        stats_state = {"header_written": False}
        if single_sheet:
            single_ws = wb.create_sheet(title='All Pages')

        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            start = max(1, start_page)
            end = total_pages if end_page is None else min(end_page, total_pages)
            for page_num in range(start, end + 1):
                page = pdf.pages[page_num - 1]
                invoice_table = self._extract_invoice_table(page)
                if invoice_table:
                    headers, rows = invoice_table
                    if single_sheet:
                        if not invoice_header_written:
                            single_ws.append(headers)
                            invoice_header_written = True
                        for row in rows:
                            single_ws.append(row)
                        single_ws.append([])
                    else:
                        sheet_name = f'Page {page_num} - Invoice'
                        ws = wb.create_sheet(title=sheet_name[:31])
                        ws.append(headers)
                        for row in rows:
                            ws.append(row)

                    if progress_cb:
                        progress_cb(page_num, total_pages)
                    continue

                line_settings = {
                    'vertical_strategy': 'lines',
                    'horizontal_strategy': 'lines',
                    'intersection_tolerance': 5,
                    'snap_tolerance': 3,
                    'join_tolerance': 3,
                    'edge_min_length': 3,
                }
                text_settings = {
                    'vertical_strategy': 'text',
                    'horizontal_strategy': 'text',
                    'intersection_tolerance': 5,
                    'snap_tolerance': 3,
                    'join_tolerance': 3,
                }

                tables = page.extract_tables(line_settings)
                tables = [t for t in tables if t and any(any(cell for cell in row) for row in t)]
                if not tables:
                    tables = page.extract_tables(text_settings)
                    tables = [t for t in tables if t and any(any(cell for cell in row) for row in t)]

                if tables:
                    for table_index, table in enumerate(tables, start=1):
                        cleaned_rows = []
                        max_cols = 0
                        for row in table:
                            cleaned = [cell if cell is not None else '' for cell in row]
                            if any(str(cell).strip() for cell in cleaned):
                                cleaned_rows.append(cleaned)
                                max_cols = max(max_cols, len(cleaned))

                        cleaned_rows = self._normalize_stats_rows(cleaned_rows, stats_state)

                        if single_sheet:
                            for row in cleaned_rows:
                                if len(row) < max_cols:
                                    row.extend([''] * (max_cols - len(row)))
                                single_ws.append(row)
                            single_ws.append([])
                        else:
                            sheet_name = f'Page {page_num} - Table {table_index}'
                            ws = wb.create_sheet(title=sheet_name[:31])
                            for row in cleaned_rows:
                                if len(row) < max_cols:
                                    row.extend([''] * (max_cols - len(row)))
                                ws.append(row)
                else:
                    text = page.extract_text() or ''
                    if single_sheet:
                        for line in text.splitlines():
                            single_ws.append([line])
                        single_ws.append([])
                    else:
                        sheet_name = f'Page {page_num}'
                        ws = wb.create_sheet(title=sheet_name[:31])
                        for line in text.splitlines():
                            ws.append([line])

                if progress_cb:
                    progress_cb(page_num, total_pages)

        if not wb.sheetnames:
            wb.create_sheet('Sheet1')

        wb.save(excel_path)

    def _extract_invoice_table(self, page):
        words = page.extract_words()
        if not words:
            return None

        amount_pattern = re.compile(r'^-?\d{1,3}(?:,\d{3})*(?:\.\d{2})$|^-?\d+\.\d{2}$')
        lines = self._group_words_by_line(words)
        header_index = None
        header_line = None
        for idx, line in enumerate(lines):
            texts = {w['text'] for w in line}
            if {'ITEM/DESCRIPTION', 'AMOUNT'}.issubset(texts):
                header_index = idx
                header_line = line
                break

        if header_line is None:
            return None

        columns = ['Sold', 'B/O', 'Del', 'ITEM/DESCRIPTION', 'AMOUNT']
        positions = {}
        for w in header_line:
            if w['text'] in columns:
                positions[w['text']] = (w['x0'], w['x1'])

        if not all(col in positions for col in columns):
            return None

        boundaries = [-float('inf')]
        for current_col, next_col in zip(columns, columns[1:]):
            current_x0, current_x1 = positions[current_col]
            next_x0, _ = positions[next_col]
            boundaries.append((current_x1 + next_x0) / 2)
        boundaries.append(float('inf'))

        rows = []
        for line in lines[header_index + 1:]:
            line_text = ' '.join(w['text'] for w in line).strip()
            if not line_text:
                continue
            if set(line_text) == {'-'}:
                continue
            if 'CONTINUATION ON NEXT PAGE' in line_text:
                continue

            cells = [''] * len(columns)
            amount_word = None
            for w in reversed(line):
                if amount_pattern.match(w['text']):
                    amount_word = w
                    break

            for w in line:
                if amount_word is not None and w is amount_word:
                    continue
                center = (w['x0'] + w['x1']) / 2
                col_index = next(
                    i for i in range(len(columns))
                    if boundaries[i] <= center < boundaries[i + 1]
                )
                if col_index == len(columns) - 1 and not amount_pattern.match(w['text']):
                    col_index = 3
                cells[col_index] = f"{cells[col_index]} {w['text']}".strip()

            if amount_word is not None:
                cells[-1] = amount_word['text']

            has_amount = bool(cells[-1])
            has_quantities = any(cells[:3])
            if amount_word is None and not has_quantities and rows:
                rows[-1][3] = f"{rows[-1][3]} {line_text}".strip()
                continue

            if any(cells):
                rows.append(cells)

        if rows and not any(row[1] for row in rows) and any(row[2] for row in rows):
            for row in rows:
                if row[2] and not row[1]:
                    row[1] = row[2]
                    row[2] = ''

        if not rows:
            return None

        return (columns, rows)

    def _group_words_by_line(self, words, tolerance=3):
        words = sorted(words, key=lambda w: (w['top'], w['x0']))
        lines = []
        current = []
        current_top = None

        for word in words:
            if current_top is None or abs(word['top'] - current_top) <= tolerance:
                current.append(word)
                if current_top is None:
                    current_top = word['top']
            else:
                lines.append(current)
                current = [word]
                current_top = word['top']

        if current:
            lines.append(current)

        return [sorted(line, key=lambda w: w['x0']) for line in lines]

    def _normalize_stats_rows(self, rows, state=None):
        if state is None:
            state = {"header_written": False}
        day_regex = re.compile(r'\b(mon|tue|wed|thu|fri|sat|sun)\w*', re.IGNORECASE)
        month_regex = re.compile(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)\w*', re.IGNORECASE)
        year_regex = re.compile(r'\b20\d{2}\b')

        day_tokens = {'monday','tuesday','wednesday','thursday','friday','saturday','sunday'}
        month_tokens = {'january','february','march','april','may','june','july','august','september','october','november','december'}

        normalized = []
        numeric_pattern = re.compile(r'^-?\d{1,3}(?:,\d{3})*(?:\.\d+)?$|^-?\d+\.\d+$')
        for row in rows:
            cells = [str(cell).strip() if cell is not None else '' for cell in row]
            if not any(cells):
                continue

            has_pipe = any('|' in cell for cell in cells)
            if has_pipe:
                cells = self._align_pipe_column(cells, target_index=7)

            joined = ' '.join(cells)
            lower_joined = joined.lower()

            if not has_pipe and not any(day in lower_joined for day in day_tokens) and not any(month in lower_joined for month in month_tokens):
                if len(cells) < 12:
                    cells.extend([''] * (12 - len(cells)))
                elif len(cells) > 12:
                    cells = cells[:12]
                normalized.append(cells)
                continue

            raw_text = joined
            compact_text = re.sub(r'[^A-Za-z0-9]', '', raw_text).lower()
            has_day = bool(day_regex.search(raw_text)) or any(d in compact_text for d in day_tokens)
            has_month = bool(month_regex.search(raw_text)) or any(m in compact_text for m in month_tokens)
            has_year = bool(year_regex.search(raw_text)) or '20' in compact_text

            is_header = (
                'sales' in compact_text and 'profit' in compact_text and
                'qte' in compact_text and 'sbill' in compact_text and
                'average' in compact_text
            )
            is_separator = set(compact_text) in ({''}, {'-'}, {'='})

            if is_header:
                if not state.get("header_written"):
                    normalized.append([
                        'DATE', '', '', '',
                        'SALES', 'PROFIT', '(%)', '|', 'QTE', 'S/BILL', 'AVERAGE', '(%)'
                    ])
                    state["header_written"] = True
                continue

            if is_separator:
                continue

            if has_day and has_month and has_year:
                try:
                    sep_index = next(i for i, cell in enumerate(cells) if '|' in cell)
                except StopIteration:
                    sep_index = None

                if sep_index is not None:
                    left_tokens = [cell for cell in cells[:sep_index] if cell and cell != '|']
                    date_text = ' '.join(left_tokens).replace('  ', ' ').strip()
                    for i in range(len(cells)):
                        cells[i] = ''
                    cells[0] = date_text
                    cells[sep_index] = '|'
                else:
                    date_text = ' '.join([cell for cell in cells if cell]).replace('  ', ' ').strip()
                    cells = [''] * len(cells)
                    cells[0] = date_text

            if '|' in cells:
                cells = self._normalize_stats_columns(cells, numeric_pattern)

            if len(cells) < 12:
                cells.extend([''] * (12 - len(cells)))
            if len(cells) > 12:
                cells = cells[:12]

            normalized.append(cells)

        return normalized

    def _normalize_stats_columns(self, cells, numeric_pattern):
        if '|' not in cells:
            return cells

        pipe_index = cells.index('|')
        left = cells[:pipe_index]

        lower_left = ' '.join(left).lower()
        if any(day in lower_left for day in ['monday','tuesday','wednesday','thursday','friday','saturday','sunday']):
            return cells

        code = None
        for token in left:
            if token and not numeric_pattern.match(token):
                code = token
                break

        nums = [token for token in left if token and numeric_pattern.match(token)]

        new_left = [''] * pipe_index
        if code:
            new_left[0] = code

        if nums:
            for offset, val in enumerate(nums[:3]):
                target = 4 + offset
                if target < pipe_index:
                    new_left[target] = val

        return new_left + cells[pipe_index:]

    def _align_pipe_column(self, cells, target_index=7):
        if not cells:
            return cells

        pipe_index = None
        right_text = None
        for i, cell in enumerate(cells):
            if '|' in cell:
                pipe_index = i
                if cell != '|':
                    left, right = cell.split('|', 1)
                    cells[i] = left.strip()
                    right_text = right.strip()
                break

        if pipe_index is None:
            return cells

        if pipe_index < target_index:
            insert_count = target_index - pipe_index
            cells = cells[:pipe_index] + ([''] * insert_count) + cells[pipe_index:]
        elif pipe_index > target_index:
            remove_count = pipe_index - target_index
            left = cells[:pipe_index]
            right = cells[pipe_index:]
            while remove_count > 0 and left:
                if left[-1] == '':
                    left.pop()
                    remove_count -= 1
                else:
                    break
            cells = left + right

        if len(cells) <= target_index:
            cells.extend([''] * (target_index + 1 - len(cells)))

        cells[target_index] = '|'
        if right_text:
            if len(cells) <= target_index + 1:
                cells.append(right_text)
            elif cells[target_index + 1] == '':
                cells[target_index + 1] = right_text

        return cells


def convert_range_worker(args):
    pdf_path, excel_path, start_page, end_page, single_sheet = args
    ConverterService().convert_range(pdf_path, excel_path, start_page, end_page, single_sheet=single_sheet)
    return excel_path


def merge_workbooks(source_paths, output_path, single_sheet: bool = False):
    merged = Workbook()
    merged.remove(merged.active)

    if single_sheet:
        ws = merged.create_sheet(title='All Pages')
        for source_path in source_paths:
            wb = load_workbook(source_path)
            if 'All Pages' in wb.sheetnames:
                sheet = wb['All Pages']
                for row in sheet.iter_rows(values_only=True):
                    ws.append(list(row))
            else:
                for sheet in wb.worksheets:
                    ws.append([f"[{sheet.title}]"])
                    for row in sheet.iter_rows(values_only=True):
                        ws.append(list(row))
            ws.append([])
    else:
        for source_path in source_paths:
            wb = load_workbook(source_path)
            for sheet in wb.worksheets:
                title = sheet.title
                if title in merged.sheetnames:
                    suffix = 1
                    while f"{title[:28]}_{suffix}" in merged.sheetnames:
                        suffix += 1
                    title = f"{title[:28]}_{suffix}"
                ws = merged.create_sheet(title=title[:31])
                for row in sheet.iter_rows(values_only=True):
                    ws.append(list(row))

    if not merged.sheetnames:
        merged.create_sheet('Sheet1')

    merged.save(output_path)